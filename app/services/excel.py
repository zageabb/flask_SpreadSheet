"""Excel workbook interchange with practical formatting preservation."""
from __future__ import annotations

import io
import json
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook as ExcelWorkbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session, select

from ..models import Sheet, SheetCell, Workbook


class ExcelWorkbookService:
    def __init__(self, session: Session) -> None:
        self.session = session

    def export_workbook(self, workbook_id: int) -> tuple[str, io.BytesIO]:
        workbook = self.session.get(Workbook, workbook_id)
        if workbook is None:
            raise LookupError("Workbook not found")
        target = ExcelWorkbook()
        target.remove(target.active)
        sheets = self.session.exec(select(Sheet).where(Sheet.workbook_id == workbook_id).order_by(Sheet.id)).all()
        used: set[str] = set()
        for sheet in sheets:
            title = _unique_sheet_name(sheet.name, used)
            used.add(title)
            output_sheet = target.create_sheet(title)
            cells = self.session.exec(select(SheetCell).where(SheetCell.sheet_id == sheet.id)).all()
            for cell in cells:
                target_cell = output_sheet.cell(row=cell.row_index + 1, column=cell.col_index + 1)
                target_cell.value = cell.formula or cell.value
                if cell.number_format:
                    target_cell.number_format = cell.number_format
                _apply_style(target_cell, _load_style(cell.style_json))
        if not sheets:
            target.create_sheet("Sheet 1")
        buffer = io.BytesIO()
        target.save(buffer)
        buffer.seek(0)
        return workbook.name, buffer

    def import_workbook(self, stream: BinaryIO, source_name: str) -> int:
        source = load_workbook(stream, data_only=False)
        workbook = Workbook(name=Path(source_name).stem or "Imported workbook")
        self.session.add(workbook)
        self.session.flush()
        used: set[str] = set()
        for source_sheet in source.worksheets:
            name = _unique_sheet_name(source_sheet.title, used)
            used.add(name)
            sheet = Sheet(
                workbook_id=workbook.id,
                name=name,
                row_count=max(1, source_sheet.max_row or 1),
                col_count=max(1, source_sheet.max_column or 1),
            )
            self.session.add(sheet)
            self.session.flush()
            for row in source_sheet.iter_rows():
                for source_cell in row:
                    style = _extract_style(source_cell)
                    if source_cell.value is None and not style:
                        continue
                    raw = "" if source_cell.value is None else str(source_cell.value)
                    formula = raw if raw.startswith("=") else None
                    cell = SheetCell(
                        sheet_id=sheet.id,
                        row_index=source_cell.row - 1,
                        col_index=source_cell.column - 1,
                        value=raw or None,
                        formula=formula,
                        value_type="formula" if formula else _excel_value_type(source_cell.data_type),
                        number_format=source_cell.number_format if source_cell.number_format != "General" else None,
                        style_json=json.dumps(style, separators=(",", ":")),
                    )
                    self.session.add(cell)
        if not source.worksheets:
            self.session.add(Sheet(workbook_id=workbook.id, name="Sheet 1", row_count=12, col_count=8))
        self.session.commit()
        return workbook.id


def _unique_sheet_name(name: str, used: set[str]) -> str:
    sanitized = "".join("_" if char in "[]:*?/\\" else char for char in (name or "Sheet")).strip()[:31] or "Sheet"
    candidate = sanitized
    counter = 2
    while candidate in used:
        suffix = f" ({counter})"
        candidate = f"{sanitized[:31-len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _excel_value_type(data_type: str) -> str:
    return {"n": "number", "b": "boolean", "d": "date"}.get(data_type, "text")


def _extract_style(cell) -> dict:
    style: dict[str, object] = {}
    if cell.font:
        if cell.font.bold: style["bold"] = True
        if cell.font.italic: style["italic"] = True
        if cell.font.underline and cell.font.underline != "none": style["underline"] = True
        if cell.font.color and cell.font.color.type == "rgb": style["fontColor"] = cell.font.color.rgb
    if cell.fill and cell.fill.fill_type == "solid" and cell.fill.fgColor.type == "rgb":
        style["fillColor"] = cell.fill.fgColor.rgb
    if cell.alignment:
        if cell.alignment.horizontal: style["horizontalAlign"] = cell.alignment.horizontal
        if cell.alignment.vertical: style["verticalAlign"] = cell.alignment.vertical
        if cell.alignment.wrap_text: style["wrapText"] = True
    return style


def _load_style(raw: str | None) -> dict:
    try:
        value = json.loads(raw or "{}")
        return value if isinstance(value, dict) else {}
    except json.JSONDecodeError:
        return {}


def _apply_style(cell, style: dict) -> None:
    if any(key in style for key in ("bold", "italic", "underline", "fontColor")):
        cell.font = Font(bold=bool(style.get("bold")), italic=bool(style.get("italic")), underline="single" if style.get("underline") else None, color=style.get("fontColor"))
    if style.get("fillColor"):
        cell.fill = PatternFill(fill_type="solid", fgColor=style["fillColor"])
    if any(key in style for key in ("horizontalAlign", "verticalAlign", "wrapText")):
        cell.alignment = Alignment(horizontal=style.get("horizontalAlign"), vertical=style.get("verticalAlign"), wrap_text=bool(style.get("wrapText")))
