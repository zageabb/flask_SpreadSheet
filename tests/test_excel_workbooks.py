import io
from copy import copy

from openpyxl import Workbook, load_workbook


def _sample_workbook() -> io.BytesIO:
    source = Workbook()
    first = source.active
    first.title = "Prices"
    first["A1"] = "Cost"
    bold_font = copy(first["A1"].font)
    bold_font.bold = True
    first["A1"].font = bold_font
    first["A2"] = 12.5
    first["A2"].number_format = "£#,##0.00"
    second = source.create_sheet("Summary")
    second["A1"] = "=Prices!A2"
    output = io.BytesIO()
    source.save(output)
    output.seek(0)
    return output


def test_multisheet_excel_import_and_export(client):
    response = client.post(
        "/api/workbooks/import.xlsx",
        data={"file": (_sample_workbook(), "supplier-prices.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 201
    workbook_id = response.get_json()["workbookId"]

    exported = client.get(f"/api/workbooks/{workbook_id}/export.xlsx")
    assert exported.status_code == 200
    result = load_workbook(io.BytesIO(exported.data), data_only=False)
    assert result.sheetnames == ["Prices", "Summary"]
    assert result["Prices"]["A1"].font.bold is True
    assert result["Prices"]["A2"].number_format == "£#,##0.00"
    assert result["Summary"]["A1"].value == "=Prices!A2"


def test_cell_formatting_api_persists_style(client):
    sheet_id = client.get("/api/grid").get_json()["sheetId"]
    response = client.patch(
        f"/api/sheets/{sheet_id}/formatting",
        json={"cells": [{"row": 1, "col": 2}], "style": {"bold": True, "horizontalAlign": "right"}, "numberFormat": "0.00"},
    )
    assert response.status_code == 200
    result = client.get(f"/api/sheets/{sheet_id}/formatting").get_json()["cells"]
    assert result == [{"row": 1, "col": 2, "style": {"bold": True, "horizontalAlign": "right"}, "numberFormat": "0.00"}]
